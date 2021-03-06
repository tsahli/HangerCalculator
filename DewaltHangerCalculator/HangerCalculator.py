#!python3
# HangerCalculator.py
# Calculates hanger data and prepares label for mail merge

import openpyxl, pprint, time
from openpyxl.styles import Alignment
from fractions import Fraction

# This function is needed to sum allthread and strut mixed number strings to floats
def convert_to_float(frac_str):
    try:
        return float(frac_str)
    except ValueError:
        num, denom = frac_str.split('/')
        try:
            leading, num = num.split(' ')
            whole = float(leading)
        except ValueError:
            whole = 0
        frac = float(num) / float(denom)
        return whole - frac if whole < 0 else whole + frac

while True:
    try:
        print("\nReadMe:\nColumn A: Site Area\nColumn B: Hanger ID\nColumn C: Attachment 1 Elevation\nColumn D: Material\nColumn E: Support Span\nColumn F: Support 1 Cut Length\n\n")
        excelSheetName = input('Enter Name of Excel Sheet: ')
        jobNumber = input('Enter the job number: ')
        jobName = input('Enter the job name: ')
        wb = openpyxl.load_workbook(excelSheetName + '.xlsx')
        sheet = wb.get_sheet_by_name(excelSheetName)
        break
    except:
        print("\n--------------------------------------\n\nSomething went wrong.\nCheck workbook spelling. Your input is case sensitive.\nEnsure first sheet is the same name as the workbook.\nEnsure this program is in the same folder as the workbook.\n\n--------------------------------------")


createdSheet = wb.create_sheet()
createdSheet.title = 'Total Strut'
createdAllthreadSheet = wb.create_sheet()
createdAllthreadSheet.title = 'Total Allthread Cuts'
createdAssemblySheet = wb.create_sheet()
createdAssemblySheet.title = 'Total Assemblies'
createdConcatSheet = wb.create_sheet()
createdConcatSheet.title = 'Print Me'

hangerListTotal = []
hangerList = []
strutLengthList = []
strutLengthListTotal = []
strutTypeList = []
strutTypeListTotal = []
allthreadList = []
allthreadListTotal = []
allthreadLengthList = []
strutTypeLengthList = []
strutNameList = []
next_strut_length_row = 4
next_assembly_row = 4
next_allthread_row = 4
next_concat_row = 2

# below creates title blocks for each created sheet
createdSheet.merge_cells('A1:B2')
title = createdSheet.cell(row = 1, column = 1)
title.value = 'Total Strut'
title.alignment = Alignment(horizontal = 'center', vertical = 'center')
createdSheet.cell(row = 3, column = 1).value = "Strut Type"
createdSheet.cell(row = 3, column = 2).value = "Quantity"

createdAllthreadSheet.merge_cells('A1:B2')
title = createdAllthreadSheet.cell(row = 1, column = 1)
title.value = 'Total Allthread'
title.alignment = Alignment(horizontal = 'center', vertical = 'center')
createdAllthreadSheet.cell(row = 3, column = 1).value = "Allthread Length"
createdAllthreadSheet.cell(row = 3, column = 2).value = "Quantity"

createdAssemblySheet.merge_cells('A1:B2')
title = createdAssemblySheet.cell(row = 1, column = 1)
title.value = 'Assembly Name and Quantity'
title.alignment = Alignment(horizontal = 'center', vertical = 'center')
createdAssemblySheet.cell(row = 3, column = 1).value = "Assembly Name"
createdAssemblySheet.cell(row = 3, column = 2).value = "Quantity"

createdConcatSheet.cell(row = 1, column = 1).value = "PRINT_ME"

# Below generates 2 lists of values
for row in range(4, sheet.max_row+1):
    try:
       hangerID = sheet['B' + str(row)].value
       hangerList.append(hangerID)

       strutType = sheet['D' + str(row)].value + ': ' + sheet['E' + str(row)].value
       strutTypeList.append(strutType)

       allthreadLength = sheet['F' + str(row)].value
       allthreadList.append(allthreadLength)

       topOfStrut = sheet['C' + str(row)].value
       areaName = sheet['A' + str(row)].value

       if allthreadLength not in allthreadListTotal:
           allthreadListTotal.append(allthreadLength)

       if strutType not in strutTypeListTotal:
           strutTypeListTotal.append(strutType)

       if hangerID not in hangerListTotal:
           hangerListTotal.append(hangerID)
    except:
        continue

    # Below creates the concatenation to be used in mail merge
    label = areaName + " Tag: " + hangerID + "                                                           TOU: " + topOfStrut + "                                                                    " + strutType + "                                                               Allthread Length: " + allthreadLength + "                                                               " + jobNumber + "-" + jobName
    createdConcatSheet.cell(column = 1, row = next_concat_row, value = label)
    next_concat_row += 1

allthreadListTotal.sort()
strutTypeListTotal.sort()
hangerListTotal.sort()

# Below prints total counts on each created sheet
for x in hangerListTotal:
    createdAssemblySheet.cell(column = 1, row = next_assembly_row, value = x)
    createdAssemblySheet.cell(column = 2, row = next_assembly_row, value = hangerList.count(x))
    next_assembly_row += 1

for x in strutTypeListTotal:
    createdSheet.cell(column = 1, row = next_strut_length_row, value = x)
    createdSheet.cell(column = 2, row = next_strut_length_row, value = strutTypeList.count(x))
    next_strut_length_row += 1

for x in allthreadListTotal:
    createdAllthreadSheet.cell(column = 1, row = next_allthread_row, value = x)
    createdAllthreadSheet.cell(column = 2, row = next_allthread_row, value = (allthreadList.count(x) * 2))
    next_allthread_row += 1

# Creating sum of allthread length
for x in allthreadList:
    x = x.replace('"','')
    x = convert_to_float(x)
    allthreadLengthList.append(x)

# Creating sum of strut length
for x in strutTypeList:
    x = x.split(': ')
    x = x[1]
    x = x.replace('"','')
    x = convert_to_float(x)
    strutTypeLengthList.append(x)

# Summing lists and printing in feet
totalAllthreadLength = sum(allthreadLengthList) * 2
totalAlltrheadLengthFeet = totalAllthreadLength / 12
totalStrutLength = sum(strutTypeLengthList)
totalStrutLengthFeet = totalStrutLength / 12
totalAlltrheadLengthFeet = round(totalAlltrheadLengthFeet, 2)
totalStrutLengthFeet = round(totalStrutLengthFeet, 2)
createdAllthreadSheet.cell(column = 1, row = next_allthread_row, value = "Total allthread length = " + str(totalAlltrheadLengthFeet) + " ft")
createdSheet.cell(column = 1, row = next_strut_length_row, value = "Total strut length = " + str(totalStrutLengthFeet) + " ft")

wb.save(excelSheetName + '.xlsx')
print('------------------------Done------------------------')
time.sleep(3)